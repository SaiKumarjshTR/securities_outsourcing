"""
Excel → SGML Converter  (v4.0 Pure-Rules)
Single-file entry point: python excel_batch_converter.py input.xlsx [-o output.sgm]
"""

import openpyxl
from datetime import datetime
import os
import re
import sys
sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)


# ─── CONVERTER CLASS (v4.0 Pure-Rules, extracted from notebook) ──────────────
class EnhancedExcelToSGMLConverter:
    """
    Enhanced Excel to SGML Converter v4.0 – Pure Rule-Based.
    Three critical fixes:
      1. Strict bold checking  (font.bold is True)
      2. Better empty detection (strip whitespace)
      3. Individual cell checking in metadata table
    """

    def __init__(self):
        self.row_num = 0

    # ── entity / formatting helpers ──────────────────────────────────────────

    def convert_entities(self, text):
        if not text:
            return ''
        text = str(text)
        text = text.replace('&', '&amp;')
        text = text.replace('<', '&lt;')
        text = text.replace('>', '&gt;')
        # smart quotes
        text = text.replace('\u201c', '&ldquo;')
        text = text.replace('\u201d', '&rdquo;')
        text = text.replace('\u2018', '&lsquo;')
        text = text.replace('\u2019', '&rsquo;')
        text = text.replace('\u2014', '&mdash;')
        text = text.replace('\u2013', '&ndash;')
        # French/accented characters (D6-entity fix)
        text = text.replace('\u00e9', '&eacute;')   # é
        text = text.replace('\u00e8', '&egrave;')    # è
        text = text.replace('\u00ea', '&ecirc;')     # ê
        text = text.replace('\u00e0', '&agrave;')    # à
        text = text.replace('\u00e2', '&acirc;')     # â
        text = text.replace('\u00f4', '&ocirc;')     # ô
        text = text.replace('\u00ee', '&icirc;')     # î
        text = text.replace('\u00e7', '&ccedil;')    # ç
        text = text.replace('\u00f9', '&ugrave;')    # ù
        text = text.replace('\u00fb', '&ucirc;')     # û
        text = text.replace('\u00c9', '&Eacute;')    # É
        text = text.replace('\u00c8', '&Egrave;')    # È
        text = text.replace('\u00ca', '&Ecirc;')     # Ê
        text = text.replace('\u00c0', '&Agrave;')    # À
        text = text.replace('\u00c2', '&Acirc;')     # Â
        text = text.replace('\u00d4', '&Ocirc;')     # Ô
        text = text.replace('\u00ce', '&Icirc;')     # Î
        text = text.replace('\u00c7', '&Ccedil;')    # Ç
        text = text.replace('\u0153', '&oelig;')     # oe ligature
        text = text.replace('\u0152', '&OElig;')     # OE ligature
        return text

    def format_date(self, val):
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        return str(val)

    def format_number(self, val):
        """Format numbers: percentages as %, thousands with comma separator."""
        if val is None or val == '':
            return ''
        try:
            num = float(val)
            if 0 < num < 1 and not isinstance(val, str):
                # Percentage stored as decimal (e.g. 0.10 → 10.00%)
                return f'{num*100:.2f}%'
            if num == int(num) and abs(num) >= 1000:
                # Large integer → add commas
                return f'{int(num):,}'
            if abs(num) >= 1000:
                # Large float → add commas
                return f'{num:,.2f}'.rstrip('0').rstrip('.')
            return str(val)
        except Exception:
            return str(val)

    def is_cell_truly_bold(self, cell):
        """Fix #1: strict – only True, not None/truthy."""
        if not cell.font:
            return False
        return cell.font.bold is True

    def is_cell_empty(self, cell):
        """Fix #2: treats whitespace-only as empty."""
        if not cell.value:
            return True
        if isinstance(cell.value, str) and cell.value.strip() == '':
            return True
        return False

    def detect_doc_type(self, ws) -> int:
        """
        Detect which of the 4 Excel document types this sheet is.
        1 – Floating/tracking rates  → POLIDOC, LANDSCAPE
        2 – LSERM                    → POLIDOC, portrait
        3 – FX spot risk             → POLIDOC, portrait
        4 – Position limits          → APPENDIX, LANDSCAPE
        """
        title_raw = str(ws.cell(1, 1).value or '').lower().strip()
        first_col_headers = [str(ws.cell(r, 1).value or '').lower().strip()
                             for r in range(1, min(6, ws.max_row + 1))]
        if 'floating' in title_raw or 'tracking error margin' in title_raw:
            return 1
        if 'securities eligible for reduced margin' in title_raw or 'lserm' in title_raw:
            return 2
        if 'foreign exchange' in title_raw or 'fx spot risk' in title_raw:
            return 3
        if ('listed product type' in first_col_headers
                or 'listed product type' in title_raw
                or (not title_raw and any('listed product' in h for h in first_col_headers))):
            return 4
        # Fallback heuristics
        if not ws.cell(1, 2).value:   # title row, no 2nd col → may be LSERM
            return 2
        return 1

    def get_merged_cells_map(self, ws) -> dict:
        """
        Return dict mapping (row, col) → colspan for the top-left cell of each
        merged region.  Slave cells (not top-left) map to 0 (skip sentinel).
        Unmerged cells are NOT in the dict (get() default = 1).
        """
        merged = {}
        for mc in ws.merged_cells.ranges:
            colspan = mc.max_col - mc.min_col + 1
            for r in range(mc.min_row, mc.max_row + 1):
                for c in range(mc.min_col, mc.max_col + 1):
                    if r == mc.min_row and c == mc.min_col:
                        merged[(r, c)] = colspan  # top-left: emit with COLSPAN if > 1
                    else:
                        merged[(r, c)] = 0        # slave: skip (sentinel value)
        return merged

    def _colwds_for_type(self, doc_type: int, ncols: int, section: str) -> list:
        """
        Return list of (width_int, halign_str) tuples summing to ~100.
        section is 'meta', 'head', or 'body'.
        """
        if doc_type == 1:
            # Floating/tracking: col1 fixed at ~30, rest share the remainder
            if section == 'meta':
                return [(30, 'LEFT'), (70, 'LEFT')]
            first_w    = 30
            rest_count = max(1, ncols - 1)
            per_col    = (100 - first_w) // rest_count
            actual_first = 100 - per_col * rest_count   # absorb rounding remainder
            return [(actual_first, 'LEFT')] + [(per_col, 'RIGHT')] * rest_count

        elif doc_type == 2:
            # LSERM: col1=80 LEFT, col2=20 LEFT
            if section == 'meta':
                return [(80, 'LEFT'), (20, 'LEFT')]
            if ncols <= 2:
                return [(80, 'LEFT'), (20, 'LEFT')]
            even = 100 // ncols
            return [(80, 'LEFT')] + [(max(5, (100 - 80) // (ncols - 1)), 'LEFT')] * (ncols - 1)

        elif doc_type == 3:
            # FX spot risk: head = 4×25 CENTER; body col1=25 LEFT, rest CENTER
            if section == 'meta':
                return [(30, 'LEFT'), (70, 'LEFT')]
            if section == 'head':
                w = max(5, 100 // ncols)
                return [(w, 'CENTER')] * ncols
            return [(25, 'LEFT')] + [(max(5, 75 // max(1, ncols - 1)), 'CENTER')] * max(1, ncols - 1)

        else:  # doc_type == 4 — position limits
            FIXED = [18, 8, 26, 8, 8, 8, 8, 8, 8]
            ALIGN = ['LEFT', 'LEFT', 'LEFT', 'LEFT', 'RIGHT', 'RIGHT', 'RIGHT', 'RIGHT', 'RIGHT']
            if ncols <= len(FIXED):
                return list(zip(FIXED[:ncols], ALIGN[:ncols]))
            extra = ncols - len(FIXED)
            remaining = 100 - sum(FIXED)
            ew = max(5, remaining // extra)
            return list(zip(FIXED, ALIGN)) + [(ew, 'RIGHT')] * extra

    # ── structure analysis ───────────────────────────────────────────────────

    def analyze_sheet_structure(self, ws):
        structure = {
            'title_row':      None,
            'metadata_rows':  [],
            'header_row':     None,
            'data_start_row': None,
        }

        # Title row: row 1 has value in col-1 but empty col-2
        if ws.cell(1, 1).value and not ws.cell(1, 2).value:
            structure['title_row'] = 1

        # Metadata rows: rows 2-6 whose col-1 value contains ':'
        for row_idx in range(2, 7):
            cell_val = ws.cell(row_idx, 1).value
            if cell_val and ':' in str(cell_val):
                structure['metadata_rows'].append(row_idx)

        # Header row: first bold row after metadata
        search_start = (max(structure['metadata_rows']) + 1
                        if structure['metadata_rows'] else 2)
        for row_idx in range(search_start, search_start + 10):
            cell = ws.cell(row_idx, 1)
            if cell.value and self.is_cell_truly_bold(cell):
                structure['header_row']     = row_idx
                structure['data_start_row'] = row_idx + 1
                break

        return structure

    # ── SGML generators ──────────────────────────────────────────────────────

    def generate_metadata_table(self, ws, metadata_rows, doc_type: int = 2):
        """Metadata table with type-specific column widths."""
        n_meta_cols = min(ws.max_column, 2)
        colwds = self._colwds_for_type(doc_type, n_meta_cols, 'meta')

        sgml = []
        sgml.append('<P><TABLE>')
        sgml.append('<SGMLTBL>')
        sgml.append('<TBLBODY TBLWD="600">')
        sgml.append('<TBLCDEFS>')
        for i, (w, align) in enumerate(colwds):
            sgml.append(f'<TBLCDEF COLWD="{w}" HALIGN="{align}">')
        sgml.append('</TBLCDEFS>')
        sgml.append('<TBLROWS>')

        for row_idx in metadata_rows:
            # Skip entirely empty metadata rows (D5)
            row_vals = [ws.cell(row_idx, c).value for c in range(1, n_meta_cols + 1)]
            if all(v is None or str(v).strip() == '' for v in row_vals):
                continue
            sgml.append('<TBLROW>')
            for col_idx in range(1, n_meta_cols + 1):
                cell = ws.cell(row_idx, col_idx)
                if not self.is_cell_empty(cell):
                    raw = cell.value
                    val = self.convert_entities(self.format_date(raw) if col_idx == 2 else str(raw))
                    if self.is_cell_truly_bold(cell):
                        val = f'<BOLD>{val}</BOLD>'
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}">{val}</TBLCELL>')
                else:
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}">&nbsp;</TBLCELL>')
            sgml.append('</TBLROW>')

        sgml.append('</TBLROWS>')
        sgml.append('</TBLBODY>')
        sgml.append('</SGMLTBL>')
        sgml.append('</TABLE></P>')
        return '\n'.join(sgml)

    def generate_data_table(self, ws, header_row, data_start_row, doc_type: int = 2):
        """
        Generate data table with:
        - Type-specific column widths (D1)
        - COLSPAN for Excel merged cells (D7)
        - Skip entirely empty data rows (D5)
        - Comma-formatted large numbers (D9)
        - Multi-level header detection (D8)
        """
        n_cols = ws.max_column
        merged = self.get_merged_cells_map(ws)
        head_colwds = self._colwds_for_type(doc_type, n_cols, 'head')
        body_colwds = self._colwds_for_type(doc_type, n_cols, 'body')

        sgml = []
        sgml.append('<P><TABLE>')
        sgml.append('<SGMLTBL>')

        # ── Table HEAD ─────────────────────────────────────────────────────
        # Detect multi-level header (D8): if row above header_row is also bold & non-empty
        multi_header_start = header_row
        if header_row > 1:
            prev = ws.cell(header_row - 1, 1)
            if prev.value and self.is_cell_truly_bold(prev):
                multi_header_start = header_row - 1

        sgml.append('<TBLHEAD TBLWD="600">')
        sgml.append('<TBLCDEFS TOPSEP="HSINGLE" COLSEP="VSINGLE">')
        for w, align in head_colwds:
            sgml.append(f'<TBLCDEF COLWD="{w}" HALIGN="{align}">')
        sgml.append('</TBLCDEFS>')
        sgml.append('<TBLROWS><?TBLROW 1>')
        for h_row in range(multi_header_start, header_row + 1):
            is_last = (h_row == header_row)
            row_attr = ' ROWSEP="HSINGLE"' if is_last else ''
            sgml.append(f'<TBLROW{row_attr}>')
            col_idx = 1
            while col_idx <= n_cols:
                cell = ws.cell(h_row, col_idx)
                m_state = merged.get((h_row, col_idx), 1)  # 1=unmerged, 0=slave, N=colspan
                if m_state == 0:  # slave cell — skip
                    col_idx += 1
                    continue
                colspan_attr = f' COLSPAN="{m_state}"' if m_state > 1 else ''
                if not self.is_cell_empty(cell):
                    val = self.convert_entities(str(cell.value))
                    if self.is_cell_truly_bold(cell):
                        val = f'<BOLD>{val}</BOLD>'
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}"{colspan_attr}>{val}</TBLCELL>')
                else:
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}"{colspan_attr}>&nbsp;</TBLCELL>')
                col_idx += 1
            sgml.append('</TBLROW>')
        sgml.append('</TBLROWS>')
        sgml.append('</TBLHEAD>')

        # ── Table BODY ─────────────────────────────────────────────────────
        sgml.append('<TBLBODY TBLWD="600">')
        sgml.append('<TBLCDEFS>')
        for w, align in body_colwds:
            sgml.append(f'<TBLCDEF COLWD="{w}" HALIGN="{align}">')
        sgml.append('</TBLCDEFS>')
        sgml.append('<TBLROWS><?TBLROW 2>')
        row_counter = 2

        for row_idx in range(data_start_row, ws.max_row + 1):
            # D5: skip entirely empty data rows
            row_vals = [ws.cell(row_idx, c).value for c in range(1, n_cols + 1)]
            if all(v is None or str(v).strip() == '' for v in row_vals):
                continue

            row_counter += 1
            sgml.append(f'<?TBLROW {row_counter}>')
            sgml.append('<TBLROW>')
            col_idx = 1
            while col_idx <= n_cols:
                cell = ws.cell(row_idx, col_idx)
                m_state = merged.get((row_idx, col_idx), 1)  # 1=unmerged, 0=slave, N=colspan
                if m_state == 0:  # slave cell — skip
                    col_idx += 1
                    continue
                colspan_attr = f' COLSPAN="{m_state}"' if m_state > 1 else ''
                if not self.is_cell_empty(cell):
                    # D9: format numbers with commas; D6: percentages
                    raw = cell.value
                    if isinstance(raw, (int, float)):
                        val = self.format_number(raw)
                    else:
                        val = str(raw)
                    val = self.convert_entities(val)
                    if self.is_cell_truly_bold(cell):
                        val = f'<BOLD>{val}</BOLD>'
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}"{colspan_attr}>{val}</TBLCELL>')
                else:
                    sgml.append(f'<TBLCELL COLSTART="{col_idx}"{colspan_attr}>&nbsp;</TBLCELL>')
                col_idx += 1
            sgml.append('</TBLROW>')

        sgml.append('</TBLROWS>')
        sgml.append('</TBLBODY>')
        sgml.append('</SGMLTBL>')
        sgml.append('</TABLE></P>')
        return '\n'.join(sgml)

    # ── main convert ─────────────────────────────────────────────────────────

    def convert(self, input_excel, output_sgml, doc_metadata):
        """
        Full conversion with all business fixes applied:
        E3: No INITID; D2: LANDSCAP; D3: APPENDIX root; D1/D5/D7/D8/D9 in table generators.
        """
        print(f'\nLoading: {os.path.basename(input_excel)}')
        wb = openpyxl.load_workbook(input_excel)
        ws = wb.active
        print(f'  Rows: {ws.max_row}  Cols: {ws.max_column}')

        doc_type  = self.detect_doc_type(ws)
        landscape = doc_type in (1, 4)          # D2
        root_tag  = 'APPENDIX' if doc_type == 4 else 'POLIDOC'  # D3
        print(f'  Doc type: {doc_type}  Root: <{root_tag}>  Landscape: {landscape}')

        structure = self.analyze_sheet_structure(ws)
        print(f'  Title row:    {structure["title_row"]}')
        print(f'  Metadata rows:{structure["metadata_rows"]}')
        print(f'  Header row:   {structure["header_row"]}')
        print(f'  Data starts:  {structure["data_start_row"]}')

        lines = []

        # ── Root tag (E3: no INITID; D3: APPENDIX vs POLIDOC) ────────────────
        adddate  = doc_metadata.get('adddate', '')
        lang     = doc_metadata.get('lang', 'EN')
        clipdate = doc_metadata.get('clipdate', '')
        moddate  = doc_metadata.get('moddate', '')
        # LABEL: per keying rules — describes the document category
        _doc_type_labels = {
            1: 'CIRO Margin Notice',
            2: 'CIRO Margin Notice',
            3: 'CIRO Margin Notice',
            4: 'Annex',
        }
        label = _doc_type_labels.get(doc_type, 'CIRO Notice')
        if root_tag == 'POLIDOC':
            lines.append(f'<POLIDOC LABEL="{label}" ADDDATE="{adddate}" LANG="{lang}" '
                         f'CLIPDATE="{clipdate}" MODDATE="{moddate}">')
            lines.append('<POLIDENT>')
            lines.append(f'<N>{self.convert_entities(doc_metadata.get("title",""))}</N>')
            lines.append(f'<DATE>{doc_metadata.get("date","")}</DATE>')
            lines.append('</POLIDENT>')
        else:
            lines.append(f'<APPENDIX LABEL="{label}" ADDDATE="{adddate}" LANG="{lang}">')

        lines.append('<FREEFORM>')

        # D2: landscape markers wrap the BLOCK2
        if landscape:
            lines.append('<?RSRVON?>')

        lines.append('<BLOCK2>')

        # Title (E2/D10: no case transformation — preserve original)
        if structure['title_row']:
            raw_title = str(ws.cell(structure['title_row'], 1).value)
            lines.append(f'<TI>{self.convert_entities(raw_title)}</TI>')

        # Metadata table
        if structure['metadata_rows']:
            lines.append(self.generate_metadata_table(ws, structure['metadata_rows'], doc_type))

        # Data table
        if structure['header_row'] and structure['data_start_row']:
            lines.append(self.generate_data_table(
                ws, structure['header_row'], structure['data_start_row'], doc_type))

        lines.append('</BLOCK2>')

        if landscape:
            lines.append('<?RSRVOFF?>')

        lines.append('</FREEFORM>')
        lines.append(f'</{root_tag}>')

        content = '\n'.join(lines)
        with open(output_sgml, 'w', encoding='utf-8') as f:
            f.write(content)

        print(f'  ✅ Saved → {os.path.basename(output_sgml)}  ({len(content):,} chars)')
        return output_sgml


# ─── VALIDATION ──────────────────────────────────────────────────────────────

def validate_sgml(sgml_path):
    with open(sgml_path, 'r', encoding='utf-8') as f:
        content = f.read()

    is_appendix = '<APPENDIX' in content
    checks = {
        'Has root tag':   ('<POLIDOC' in content or '<APPENDIX' in content),
        'Has POLIDENT':   ('<POLIDENT>' in content or is_appendix),  # APPENDIX has no POLIDENT
        'Has FREEFORM':   '<FREEFORM>'  in content,
        'Has BLOCK2':     '<BLOCK2>'    in content,
        'Has TABLE':      '<TABLE>'     in content,
        'Has TBLCELL':    '<TBLCELL'    in content,
    }

    passed = sum(checks.values())
    total  = len(checks)
    verdict = '✅ PASS' if passed == total else f'⚠️  {passed}/{total}'
    details = '  '.join(f'{"✅" if v else "❌"} {k}' for k, v in checks.items())
    return verdict, details, content.count('<TBLCELL'), content.count('<BOLD>')


# ─── BATCH RUNNER ────────────────────────────────────────────────────────────

def _parse_excel_date_to_str(val):
    """Convert an Excel cell date value to 'Month D, YYYY' string (no leading zero)."""
    if val is None:
        return None
    # datetime / date object from openpyxl
    if hasattr(val, 'year'):
        return f'{val.strftime("%B")} {val.day}, {val.year}'
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
            try:
                dt = datetime.strptime(val, fmt)
                return f'{dt.strftime("%B")} {dt.day}, {dt.year}'
            except ValueError:
                pass
    return None


def derive_metadata(input_excel):
    """
    Build a doc_metadata dict from filename + Excel content.
    - Title  : read from cell A1
    - Date   : read from row 4 col 2 (Effective date), fallback row 3 col 2 (Release date),
               fallback to today
    - ADDDATE/CLIPDATE/MODDATE : today in YYYYMMDD
    - LANG   : EN  (always)
    - INITID : omitted (empty string)
    """
    stem  = os.path.splitext(os.path.basename(input_excel))[0]
    today = datetime.now().strftime('%Y%m%d')
    today_long = datetime.now().strftime('%B %-d, %Y') if os.name != 'nt' else \
                 f'{datetime.now().strftime("%B")} {datetime.now().day}, {datetime.now().year}'

    try:
        wb = openpyxl.load_workbook(input_excel, read_only=True, data_only=True)
        ws = wb.active
        title = str(ws.cell(1, 1).value or stem)
        # Row 4 col 2 = Effective date (all POLIDOC doc types)
        date_str = _parse_excel_date_to_str(ws.cell(4, 2).value)
        if not date_str:
            # Fallback: Row 3 col 2 = Release date
            date_str = _parse_excel_date_to_str(ws.cell(3, 2).value)
        if not date_str:
            date_str = today_long
        wb.close()
    except Exception:
        title    = stem
        date_str = today_long

    return {
        'title':    title,
        'date':     date_str,
        'adddate':  today,
        'lang':     'EN',
        'initid':   '',
        'clipdate': today,
        'moddate':  today,
    }


def run_batch():
    # ── Batch-mode paths (dev only — not used by single-file pipeline) ────────
    _base        = os.path.dirname(os.path.abspath(__file__))
    _output_dir  = os.path.join(_base, 'excel_sgml_output')
    os.makedirs(_output_dir, exist_ok=True)
    _input_files: list = []   # populate manually for batch runs
    # ─────────────────────────────────────────────────────────────────────────

    OUTPUT_DIR  = _output_dir
    INPUT_FILES = _input_files

    converter = EnhancedExcelToSGMLConverter()
    print('=' * 70)
    print('  Excel → SGML Batch Converter  (v4.0 Pure-Rules)')
    print(f'  Output folder: {OUTPUT_DIR}')
    print('=' * 70)

    results = []
    for xlsx_path in INPUT_FILES:
        if not os.path.exists(xlsx_path):
            print(f'\n❌ NOT FOUND: {xlsx_path}')
            results.append((os.path.basename(xlsx_path), 'FILE NOT FOUND', '', 0, 0))
            continue

        stem     = os.path.splitext(os.path.basename(xlsx_path))[0]
        out_path = os.path.join(OUTPUT_DIR, stem + '.sgm')
        meta     = derive_metadata(xlsx_path)

        try:
            converter.convert(xlsx_path, out_path, meta)
            verdict, details, cells, bolds = validate_sgml(out_path)
            results.append((os.path.basename(xlsx_path), verdict, out_path, cells, bolds))
        except Exception as exc:
            print(f'  ❌ ERROR: {exc}')
            import traceback; traceback.print_exc()
            results.append((os.path.basename(xlsx_path), f'ERROR: {exc}', '', 0, 0))

    # ── Summary ──────────────────────────────────────────────────────────────
    print('\n' + '=' * 70)
    print('  BATCH SUMMARY')
    print('=' * 70)
    for name, verdict, out_path, cells, bolds in results:
        print(f'  {verdict}  {name}')
        if out_path:
            print(f'         → {os.path.basename(out_path)}  '
                  f'({cells} cells, {bolds} bold tags)')
    print('=' * 70)
    print(f'  Output saved to: {OUTPUT_DIR}')
    print('=' * 70)


def convert_single(input_excel: str, output_sgml: str = None) -> str:
    """
    Convert a single Excel file to SGML.
    
    Args:
        input_excel : path to the input .xlsx / .xls file
        output_sgml : path for the output .sgm file (optional)
                      defaults to same folder as input, same stem + .sgm
    Returns:
        path to the generated .sgm file
    """
    input_path = os.path.abspath(input_excel)
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Input file not found: {input_path}')

    if output_sgml is None:
        stem = os.path.splitext(os.path.basename(input_path))[0]
        output_sgml = os.path.join(os.path.dirname(input_path), stem + '.sgm')

    output_path = os.path.abspath(output_sgml)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    metadata = derive_metadata(input_path)
    converter = EnhancedExcelToSGMLConverter()
    converter.convert(input_path, output_path, metadata)
    return output_path


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Excel → SGML converter (single file or batch mode)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single file (output next to input):
  python excel_batch_converter.py input.xlsx

  # Single file with explicit output path:
  python excel_batch_converter.py input.xlsx -o /path/to/output.sgm

  # Batch mode (original behaviour):
  python excel_batch_converter.py --batch
""",
    )
    parser.add_argument('input', nargs='?', help='Input Excel file (.xlsx/.xls)')
    parser.add_argument('-o', '--output', help='Output SGML file path (single-file mode only)')
    parser.add_argument('--batch', action='store_true', help='Run original batch mode')

    args = parser.parse_args()

    if args.batch or args.input is None:
        run_batch()
    else:
        out = convert_single(args.input, args.output)
        print(f'\n✅ Output: {out}')
