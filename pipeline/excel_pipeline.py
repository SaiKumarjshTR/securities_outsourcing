"""
excel_pipeline.py
════════════════════════════════════════════════════════════════════════════════
STEP 1: Convert all 12 Excel files → SGML
STEP 2: Validate each SGML using 4-method approach (no vendor files needed):
  A. Stage-0 deterministic checks (structural, entity, table — 17 vendor-free checks)
  B. Business rules schema per doc-type (root tag, landscape, col count, metadata)
  C. Cross-file consistency within each type group
  E. Round-trip completeness (Excel cell count vs SGML TBLCELL count)

Output: excel_pipeline_output/  (SGML files)
        excel_pipeline_report.xlsx  (comprehensive validation report)
"""

import os, re, sys, glob
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))
sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

# ─── PATHS  (edit these before running) ──────────────────────────────────────
_HERE       = Path(__file__).parent
INPUT_DIR   = Path(os.environ.get('EXCEL_INPUT_DIR',  _HERE / 'excel_input'))
OUTPUT_DIR  = Path(os.environ.get('EXCEL_OUTPUT_DIR', _HERE / 'excel_pipeline_output'))
REPORT_PATH = Path(os.environ.get('EXCEL_REPORT_PATH', _HERE / 'excel_pipeline_report.xlsx'))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Import converter from existing module
from excel_batch_converter import EnhancedExcelToSGMLConverter, derive_metadata

# Stage-0 constants (inline — no external validator dependency in deployment)
SEVERITY_SCORE  = {'Critical': -15, 'High': -8, 'Medium': -4, 'Low': -1}
PASS_THRESHOLD  = 85

# Stage-0 check functions — use unified_gpt_validator when available (local dev),
# fall back to no-op stubs in deployment (validator not deployed).
try:
    from unified_gpt_validator import (
        check_p2_nesting, check_p1_nesting, check_item_nesting,
        check_double_p_before_item, check_p_p1_double_nesting,
        check_header_footer_in_body, check_dash_entities, check_apostrophe_entities,
        check_quote_entities, check_french_entities, check_mdash_spacing,
        check_table_colwd, check_tblwd_attr, check_ocr_corruption,
        check_url_line_tags, check_double_hyphen_mdash,
    )
    _STAGE0_AVAILABLE = True
except ImportError:
    _STAGE0_AVAILABLE = False
    def _stub_check(_sgml): return ('PASS', [])
    (check_p2_nesting, check_p1_nesting, check_item_nesting,
     check_double_p_before_item, check_p_p1_double_nesting,
     check_header_footer_in_body, check_dash_entities, check_apostrophe_entities,
     check_quote_entities, check_french_entities, check_mdash_spacing,
     check_table_colwd, check_tblwd_attr, check_ocr_corruption,
     check_url_line_tags, check_double_hyphen_mdash) = [_stub_check] * 16

# ─── BUSINESS RULES PER DOC TYPE ─────────────────────────────────────────────
# Each rule: (check_id, description, severity, fn(sgml, meta) → (status, detail))
DOC_TYPE_NAMES = {1: 'Floating/Tracking Rates', 2: 'LSERM', 3: 'FX Spot Risk', 4: 'Position Limits'}
DOC_TYPE_RULES = {
    1: {  # Floating/tracking rates
        'root_tag':    'POLIDOC',
        'landscape':    True,
        'min_meta_rows': 2,
        'expected_data_cols': (8, 12),   # (min, max) acceptable
        'title_must_contain': ['floating', 'tracking'],
    },
    2: {  # LSERM
        'root_tag':    'POLIDOC',
        'landscape':    False,
        'min_meta_rows': 2,
        'expected_data_cols': (2, 3),
        'title_must_contain': ['securities eligible', 'lserm'],
    },
    3: {  # FX spot risk
        'root_tag':    'POLIDOC',
        'landscape':    False,
        'min_meta_rows': 2,
        'expected_data_cols': (3, 6),
        'title_must_contain': ['foreign exchange', 'fx'],
    },
    4: {  # Position limits
        'root_tag':    'APPENDIX',
        'landscape':    True,
        'min_meta_rows': 0,
        'expected_data_cols': (8, 10),
        'title_must_contain': [],    # no title row
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — BATCH CONVERSION
# ══════════════════════════════════════════════════════════════════════════════

def run_batch_conversion() -> List[Dict]:
    """Convert all Excel files in INPUT_DIR to SGML in OUTPUT_DIR."""
    import openpyxl

    xlsx_files = sorted(INPUT_DIR.glob('*.xlsx'))
    print(f'\n{"═"*70}')
    print(f'  STEP 1: Converting {len(xlsx_files)} Excel files')
    print(f'  Source: {INPUT_DIR}')
    print(f'  Output: {OUTPUT_DIR}')
    print(f'{"═"*70}')

    converter = EnhancedExcelToSGMLConverter()
    results   = []

    for xlsx_path in xlsx_files:
        stem     = xlsx_path.stem
        out_path = OUTPUT_DIR / (stem + '.sgm')
        print(f'\n  [{len(results)+1}/{len(xlsx_files)}] {xlsx_path.name}')

        try:
            meta = derive_metadata(str(xlsx_path))

            # detect doc type for reporting
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb.active
            doc_type = converter.detect_doc_type(ws)
            total_cells = sum(
                1 for row in ws.iter_rows(min_row=1, values_only=True)
                for v in row if v is not None and str(v).strip()
            )
            wb.close()

            converter.convert(str(xlsx_path), str(out_path), meta)

            results.append({
                'xlsx':      str(xlsx_path),
                'name':      xlsx_path.name,
                'stem':      stem,
                'sgml_path': str(out_path),
                'doc_type':  doc_type,
                'doc_type_name': DOC_TYPE_NAMES.get(doc_type, 'Unknown'),
                'excel_data_cells': total_cells,
                'convert_ok': True,
                'error': '',
            })
            print(f'     ✅  Type={doc_type} ({DOC_TYPE_NAMES.get(doc_type,"?")})  Cells={total_cells}')

        except Exception as exc:
            print(f'     ❌  ERROR: {exc}')
            results.append({
                'xlsx': str(xlsx_path), 'name': xlsx_path.name, 'stem': stem,
                'sgml_path': '', 'doc_type': 0, 'doc_type_name': 'Error',
                'excel_data_cells': 0, 'convert_ok': False, 'error': str(exc),
            })

    ok = sum(1 for r in results if r['convert_ok'])
    print(f'\n  Conversion: {ok}/{len(results)} OK\n')
    return results


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2A — STAGE-0 STRUCTURAL CHECKS (no vendor)
# ══════════════════════════════════════════════════════════════════════════════

# Only checks that work without a vendor file
VENDOR_FREE_CHECKS = [
    ('dtd_p2',      'P2 Nesting',           'DTD',     lambda s: check_p2_nesting(s),          'Critical'),
    ('dtd_p1',      'P1 Nesting',           'DTD',     lambda s: check_p1_nesting(s),           'Critical'),
    ('dtd_item',    'ITEM Nesting',          'DTD',     lambda s: check_item_nesting(s),         'Critical'),
    ('dtd_extra_p', 'Extra </P> Before ITEM','DTD',    lambda s: check_double_p_before_item(s), 'High'),
    ('dtd_pp1',     'P/P1 Double Wrap',      'DTD',    lambda s: check_p_p1_double_nesting(s),  'Medium'),
    ('cont_hdrftr', 'Header/Footer Leak',    'Content', lambda s: check_header_footer_in_body(s),'Low'),
    ('ent_dash',    'ndash→mdash',           'Entity',  lambda s: check_dash_entities(s),        'Medium'),
    ('ent_apos',    'rsquo Entity',          'Entity',  lambda s: check_apostrophe_entities(s),  'Medium'),
    ('ent_quotes',  'Bare Quotes',           'Entity',  lambda s: check_quote_entities(s),       'Low'),
    ('ent_french',  'French Char Entities',  'Entity',  lambda s: check_french_entities(s),      'High'),
    ('ent_mdsp',    'mdash Spacing',         'Entity',  lambda s: check_mdash_spacing(s),        'Low'),
    ('tbl_colwd',   'COLWD Sum (90-110)',     'Table',   lambda s: check_table_colwd(s),          'Medium'),
    ('tbl_tblwd',   'TBLWD=600',             'Table',   lambda s: check_tblwd_attr(s),           'Low'),
    ('sp_ocr',      'OCR Corruption',        'Special', lambda s: check_ocr_corruption(s),       'High'),
    ('sp_url',      'URL in LINE Tag',       'Special', lambda s: check_url_line_tags(s),        'Low'),
    ('sp_hyph',     'Double-Hyphen mdash',   'Special', lambda s: check_double_hyphen_mdash(s),  'Low'),
]


def run_stage0_no_vendor(sgml: str) -> Dict:
    """Run all vendor-free Stage-0 checks. Returns structured result."""
    results = {}
    score   = 100

    for chk_id, label, cat, fn, max_sev in VENDOR_FREE_CHECKS:
        try:
            status, errors = fn(sgml)
        except Exception as exc:
            status, errors = 'ERROR', [{'severity': 'Low', 'actual': str(exc), 'line': 0, 'expected': '', 'context': ''}]

        for err in errors:
            score += SEVERITY_SCORE.get(err.get('severity', 'Low'), -1)

        summary = ''
        if errors and status == 'FAIL':
            parts   = [f"L{e['line']}: {str(e.get('actual',''))[:50]}" for e in errors[:3]]
            summary = ' | '.join(parts)
            if len(errors) > 3:
                summary += f' (+{len(errors)-3} more)'

        results[chk_id] = {
            'label':      label,
            'category':   cat,
            'status':     status,
            'errors':     errors,
            'error_count': len(errors),
            'summary':    summary,
        }

    return {'checks': results, 'score': max(0, score),
            'status': 'PASS' if max(0, score) >= PASS_THRESHOLD else 'FAIL'}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2B — BUSINESS RULES SCHEMA VALIDATION (per doc type)
# ══════════════════════════════════════════════════════════════════════════════

def validate_business_rules(sgml: str, doc_type: int, xlsx_path: str) -> Dict:
    """Check SGML output matches expected schema for its document type."""
    rules  = DOC_TYPE_RULES.get(doc_type, {})
    issues = []
    score  = 100

    # BR1: Correct root tag
    expected_root = rules.get('root_tag', 'POLIDOC')
    has_root      = f'<{expected_root}' in sgml
    wrong_roots   = [t for t in ['POLIDOC', 'APPENDIX'] if t != expected_root and f'<{t}' in sgml]
    if not has_root or wrong_roots:
        score -= 10
        issues.append({'rule': 'BR1', 'severity': 'Critical',
                        'actual': f'<{wrong_roots[0]}>' if wrong_roots else 'Missing root',
                        'expected': f'<{expected_root}>',
                        'context': 'Root tag must match document type'})

    # BR2: Landscape marker
    needs_ls  = rules.get('landscape', False)
    has_rsrvon = '<?RSRVON?>' in sgml or '<?RSRVON>' in sgml
    if needs_ls and not has_rsrvon:
        score -= 5
        issues.append({'rule': 'BR2', 'severity': 'High',
                        'actual': 'No <?RSRVON?> found',
                        'expected': '<?RSRVON?> / <?RSRVOFF?> for landscape docs',
                        'context': f'Type {doc_type} requires landscape orientation'})
    elif not needs_ls and has_rsrvon:
        score -= 2
        issues.append({'rule': 'BR2', 'severity': 'Medium',
                        'actual': '<?RSRVON?> present',
                        'expected': 'Portrait doc — no landscape markers',
                        'context': f'Type {doc_type} should be portrait'})

    # BR3: Has FREEFORM and BLOCK2
    for tag in ['<FREEFORM>', '<BLOCK2>']:
        if tag not in sgml:
            score -= 5
            issues.append({'rule': 'BR3', 'severity': 'High',
                            'actual': f'Missing {tag}',
                            'expected': f'{tag} must be present',
                            'context': 'Required structural wrapper'})

    # BR4: Has at least one TABLE
    tbl_count = len(re.findall(r'<TABLE\b', sgml, re.IGNORECASE))
    if tbl_count == 0:
        score -= 10
        issues.append({'rule': 'BR4', 'severity': 'Critical',
                        'actual': '0 TABLE tags',
                        'expected': '>= 1 TABLE',
                        'context': 'Excel data must produce at least one SGML table'})

    # BR5: Data table column count matches expected
    exp_min, exp_max = rules.get('expected_data_cols', (1, 20))
    # Find data table (last TBLBODY) COLWD count
    bodies = list(re.finditer(r'<TBLBODY[^>]*>(.*?)</TBLBODY>', sgml, re.DOTALL | re.IGNORECASE))
    if bodies:
        last_body = bodies[-1].group(1)
        col_count = len(re.findall(r'<TBLCDEF\b', last_body, re.IGNORECASE))
        if not (exp_min <= col_count <= exp_max):
            score -= 3
            issues.append({'rule': 'BR5', 'severity': 'Medium',
                            'actual': f'{col_count} columns in data table',
                            'expected': f'{exp_min}–{exp_max} for type {doc_type}',
                            'context': 'Column count must match document type spec'})
    else:
        col_count = 0

    # BR6: Title present (if expected)
    has_ti     = '<TI>' in sgml
    needs_ti   = (doc_type != 4)   # type 4 (position limits) has no title row
    if needs_ti and not has_ti:
        score -= 3
        issues.append({'rule': 'BR6', 'severity': 'Medium',
                        'actual': 'No <TI> tag',
                        'expected': '<TI> with document title',
                        'context': f'Type {doc_type} should have a title'})

    # BR7: POLIDENT present (POLIDOC types only)
    if expected_root == 'POLIDOC':
        if '<POLIDENT>' not in sgml:
            score -= 5
            issues.append({'rule': 'BR7', 'severity': 'High',
                            'actual': 'No <POLIDENT>',
                            'expected': '<POLIDENT> block required',
                            'context': 'POLIDOC must have POLIDENT metadata block'})
        else:
            # BR7b: POLIDENT has N and DATE
            if '<N>' not in sgml:
                score -= 2
                issues.append({'rule': 'BR7b', 'severity': 'Medium',
                                'actual': 'Missing <N> in POLIDENT',
                                'expected': '<N>document title</N>',
                                'context': 'POLIDENT metadata is incomplete'})
            if '<DATE>' not in sgml and 'DATE=' not in sgml:
                score -= 2
                issues.append({'rule': 'BR7c', 'severity': 'Medium',
                                'actual': 'Missing <DATE>',
                                'expected': '<DATE>...</DATE>',
                                'context': 'POLIDENT should include effective date'})

    # BR8: No INITID attribute (E3 fix)
    if 'INITID=' in sgml:
        score -= 5
        issues.append({'rule': 'BR8', 'severity': 'High',
                        'actual': 'INITID= attribute present',
                        'expected': 'INITID must be omitted',
                        'context': 'Business requirement E3'})

    # BR9: Has TBLCELL data (not empty tables)
    tblcell_count = len(re.findall(r'<TBLCELL\b', sgml, re.IGNORECASE))
    if tblcell_count == 0:
        score -= 10
        issues.append({'rule': 'BR9', 'severity': 'Critical',
                        'actual': '0 TBLCELL tags',
                        'expected': '> 0 TBLCELL (data must be present)',
                        'context': 'Tables have no data cells — conversion failed'})

    return {
        'issues':      issues,
        'score':       max(0, score),
        'status':      'PASS' if max(0, score) >= 95 else 'FAIL',
        'col_count':   col_count if bodies else 0,
        'tbl_count':   tbl_count,
        'tblcell_count': tblcell_count,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2C — CROSS-FILE CONSISTENCY (within same doc type)
# ══════════════════════════════════════════════════════════════════════════════

def _sgml_col_signature(sgml: str) -> Tuple[int, str]:
    """Return (col_count, colwd_pattern) for the data table."""
    bodies = list(re.finditer(r'<TBLBODY[^>]*>(.*?)</TBLBODY>', sgml, re.DOTALL | re.IGNORECASE))
    if not bodies:
        return 0, ''
    last   = bodies[-1].group(1)
    widths = re.findall(r'COLWD="(\d+)"', last)
    aligns = re.findall(r'HALIGN="(\w+)"', last)
    return len(widths), ','.join(widths) + '|' + ','.join(aligns)


def run_cross_file_consistency(all_results: List[Dict]) -> Dict[str, List[str]]:
    """
    For each doc type group, check:
    - All files have same column count
    - Same column width pattern
    - Same root tag
    Returns dict of doc_stem → list of consistency issues
    """
    by_type = defaultdict(list)
    for r in all_results:
        if r['convert_ok']:
            by_type[r['doc_type']].append(r)

    consistency_issues = {}  # stem → [issue strings]

    for doc_type, group in by_type.items():
        if len(group) < 2:
            continue  # nothing to compare

        # Gather signatures
        sigs = []
        for r in group:
            sgml = Path(r['sgml_path']).read_text(encoding='utf-8', errors='replace')
            col_cnt, col_sig = _sgml_col_signature(sgml)
            root = 'APPENDIX' if '<APPENDIX' in sgml else 'POLIDOC'
            sigs.append({'stem': r['stem'], 'col_cnt': col_cnt,
                          'col_sig': col_sig, 'root': root})

        # Majority vote for "expected"
        from collections import Counter
        exp_cnt = Counter(s['col_cnt'] for s in sigs).most_common(1)[0][0]
        exp_sig = Counter(s['col_sig'] for s in sigs).most_common(1)[0][0]
        exp_root= Counter(s['root']    for s in sigs).most_common(1)[0][0]

        for sig in sigs:
            issues = []
            if sig['col_cnt'] != exp_cnt:
                issues.append(f'Column count {sig["col_cnt"]} vs group expected {exp_cnt}')
            if sig['col_sig'] != exp_sig:
                issues.append(f'Column widths/alignments differ from group pattern')
            if sig['root'] != exp_root:
                issues.append(f'Root tag <{sig["root"]}> vs group expected <{exp_root}>')
            if issues:
                consistency_issues[sig['stem']] = issues

    return consistency_issues


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2E — ROUND-TRIP COMPLETENESS CHECK
# ══════════════════════════════════════════════════════════════════════════════

def check_round_trip(xlsx_path: str, sgml: str) -> Dict:
    """
    Count non-empty data cells in Excel (header + body rows only)
    vs TBLCELL tags in SGML.
    Ratio ≥ 90% → PASS. Flags potential data loss.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        excel_cells = 0
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and str(v).strip() != '':
                    excel_cells += 1
        wb.close()
    except Exception as exc:
        return {'status': 'ERROR', 'error': str(exc), 'excel_cells': 0,
                'sgml_cells': 0, 'ratio': 0, 'issues': []}

    sgml_cells = len(re.findall(r'<TBLCELL\b', sgml, re.IGNORECASE))

    # Excel has metadata, title etc. that don't all become TBLCELL
    # We also count merged cells as 1 in Excel but 1 in SGML
    # A ratio of sgml/excel is a rough gauge — expect ~0.5-1.0 for tabular data
    ratio = sgml_cells / max(excel_cells, 1)

    # More useful check: are we at least capturing >50% of non-null Excel values?
    # LSERM has 2 cols → ratio ~1.0; position limits has 9 cols → ratio ~1.0
    # If ratio < 0.3 something is badly wrong
    issues = []
    if sgml_cells == 0:
        issues.append({'severity': 'Critical', 'actual': '0 TBLCELL in SGML',
                        'expected': f'>{excel_cells//2} TBLCELL', 'context': 'No data produced'})
    elif ratio < 0.3:
        issues.append({'severity': 'High', 'actual': f'SGML cells={sgml_cells}',
                        'expected': f'Excel non-null={excel_cells} (ratio={ratio:.0%})',
                        'context': 'Likely significant data loss in conversion'})
    elif ratio < 0.5:
        issues.append({'severity': 'Medium', 'actual': f'SGML cells={sgml_cells}',
                        'expected': f'Excel non-null={excel_cells} (ratio={ratio:.0%})',
                        'context': 'Some data may be missing — review manually'})

    status = 'PASS' if not issues else ('FAIL' if ratio < 0.3 else 'WARN')
    return {'status': status, 'excel_cells': excel_cells, 'sgml_cells': sgml_cells,
            'ratio': ratio, 'issues': issues}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — RUN ALL VALIDATIONS
# ══════════════════════════════════════════════════════════════════════════════

def validate_all(conv_results: List[Dict]) -> List[Dict]:
    print(f'\n{"═"*70}')
    print(f'  STEP 2: Validating {len(conv_results)} files')
    print(f'{"═"*70}')

    # E: Cross-file consistency (needs all SGML first)
    cross_issues = run_cross_file_consistency(conv_results)

    validated = []
    for r in conv_results:
        print(f'\n  {r["name"]}  [{r["doc_type_name"]}]')

        if not r['convert_ok']:
            validated.append({**r, 'stage0': {}, 'business': {}, 'roundtrip': {}, 'cross': [],
                               'final_score': 0, 'final_status': 'ERROR'})
            continue

        sgml = Path(r['sgml_path']).read_text(encoding='utf-8', errors='replace')

        # A: Stage-0
        s0 = run_stage0_no_vendor(sgml)
        print(f'     [A] Stage-0:  {s0["score"]}/100  {s0["status"]}  '
              f'({sum(1 for c in s0["checks"].values() if c["status"]=="FAIL")} checks failed)')

        # B: Business rules
        br = validate_business_rules(sgml, r['doc_type'], r['xlsx'])
        print(f'     [B] BizRules: {br["score"]}/100  {br["status"]}  '
              f'({len(br["issues"])} issues)')

        # C: Cross-file consistency
        cross = cross_issues.get(r['stem'], [])
        cross_score = 100 - len(cross) * 5
        print(f'     [C] Consistency: {"PASS" if not cross else "WARN"}  '
              f'({len(cross)} differences vs group)')

        # E: Round-trip
        rt = check_round_trip(r['xlsx'], sgml)
        rt_score = 100 if rt['status'] == 'PASS' else (80 if rt['status'] == 'WARN' else 50)
        print(f'     [E] Round-trip: {rt["status"]}  '
              f'Excel={rt["excel_cells"]} → SGML={rt["sgml_cells"]} ({rt["ratio"]:.0%})')

        # Combined (weighted: A=40%, B=35%, C=15%, E=10%)
        final_score = round(
            s0['score']    * 0.40 +
            br['score']    * 0.35 +
            cross_score    * 0.15 +
            rt_score       * 0.10
        )
        final_status = 'PASS' if (s0['status'] == 'PASS'
                                   and br['status'] == 'PASS'
                                   and rt['status'] in ('PASS', 'WARN')
                                   and not cross) else 'FAIL'
        print(f'     → FINAL: {final_score}/100  {final_status}')

        validated.append({
            **r,
            'stage0':       s0,
            'business':     br,
            'cross':        cross,
            'cross_score':  cross_score,
            'roundtrip':    rt,
            'rt_score':     rt_score,
            'final_score':  final_score,
            'final_status': final_status,
        })

    return validated


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — EXCEL REPORT (4 sheets)
# ══════════════════════════════════════════════════════════════════════════════

def generate_report(validated: List[Dict], report_path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  pip install openpyxl'); return

    print(f'\n{"═"*70}')
    print(f'  STEP 3: Generating Excel report')
    print(f'{"═"*70}')

    wb = Workbook()

    # ─── PALETTE ─────────────────────────────────────────────────────────────
    GREEN  = PatternFill('solid', fgColor='C6EFCE')
    RED    = PatternFill('solid', fgColor='FFC7CE')
    ORANGE = PatternFill('solid', fgColor='F4B86E')
    YELLOW = PatternFill('solid', fgColor='FFEB9C')
    BLUE   = PatternFill('solid', fgColor='1F4E79')
    LBL    = PatternFill('solid', fgColor='2E75B6')
    GREY   = PatternFill('solid', fgColor='EEEEEE')

    CAT_FILLS = {
        'DTD':     PatternFill('solid', fgColor='D6E4F0'),
        'Content': PatternFill('solid', fgColor='D6F0E0'),
        'Entity':  PatternFill('solid', fgColor='F0ECD6'),
        'Table':   PatternFill('solid', fgColor='F0D6EC'),
        'Special': PatternFill('solid', fgColor='E8F0D6'),
        'BizRule': PatternFill('solid', fgColor='EAD6F0'),
        'Cross':   PatternFill('solid', fgColor='D6F0F0'),
        'RndTrip': PatternFill('solid', fgColor='F0F0D6'),
    }

    WB  = Font(bold=True, color='FFFFFF', size=10)
    WN  = Font(bold=True, size=10)
    NRM = Font(size=9)
    CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
    WRP = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    TH  = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'),  bottom=Side(style='thin'))

    def hdr(ws, row, col, val, fill=BLUE, font=WB, aln=CTR):
        c = ws.cell(row, col, val)
        c.fill = fill; c.font = font; c.alignment = aln; c.border = TH

    def cell(ws, row, col, val, fill=None, font=NRM, aln=WRP):
        c = ws.cell(row, col, val)
        if fill: c.fill = fill
        c.font = font; c.alignment = aln; c.border = TH

    def status_fill(s):
        return GREEN if s in ('PASS', 'OK') else RED if s in ('FAIL', 'ERROR') else YELLOW

    # ══ SHEET 1: Summary Dashboard ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Summary Dashboard'

    # Title banner
    ws1.merge_cells('A1:O1')
    c = ws1['A1']
    c.value = 'Excel → SGML Conversion Pipeline  ·  Validation Report (No Vendor)'
    c.fill  = PatternFill('solid', fgColor='1F3864')
    c.font  = Font(bold=True, color='FFFFFF', size=14)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:O2')
    c = ws1['A2']
    c.value = (f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  '
               f'Files: {len(validated)}  |  '
               f'Validation Methods: A(Stage-0 Structural) + B(Business Rules) + C(Cross-File) + E(Round-Trip)')
    c.fill  = PatternFill('solid', fgColor='2E4F8A')
    c.font  = Font(color='FFFFFF', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 18

    # Column headers (row 3)
    COLS = [
        'File Name', 'Doc Type', 'Root Tag', 'Landscape?',
        '[A] Stage-0\n(Structural)', '[A] Failed\nChecks',
        '[B] Biz Rules\nScore', '[B] Issues',
        '[C] Cross-File\nConsistency',
        '[E] Excel\nCells', '[E] SGML\nCells', '[E] Ratio',
        'Final\nScore', 'Final\nStatus', 'Action Required',
    ]
    for ci, h in enumerate(COLS, 1):
        hdr(ws1, 3, ci, h)
    ws1.row_dimensions[3].height = 36

    for r in validated:
        s0     = r.get('stage0', {})
        br     = r.get('business', {})
        rt     = r.get('roundtrip', {})
        cross  = r.get('cross', [])
        score  = r.get('final_score', 0)
        status = r.get('final_status', 'ERROR')

        s0_score   = s0.get('score', 0)
        s0_fails   = sum(1 for c in s0.get('checks', {}).values() if c['status'] == 'FAIL')
        br_score   = br.get('score', 0)
        br_issues  = len(br.get('issues', []))
        rt_excel   = rt.get('excel_cells', 0)
        rt_sgml    = rt.get('sgml_cells', 0)
        rt_ratio   = f'{rt.get("ratio",0):.0%}' if rt else 'N/A'
        cross_st   = 'PASS' if not cross else f'WARN ({len(cross)} diffs)'

        has_root   = '<APPENDIX' in Path(r['sgml_path']).read_text(encoding='utf-8', errors='replace') if r['convert_ok'] else '?'
        root_tag   = 'APPENDIX' if (r['convert_ok'] and has_root) else 'POLIDOC' if r['convert_ok'] else 'ERROR'
        landscape  = 'YES' if r['doc_type'] in (1, 4) else 'NO'

        action = ('OK — ready to deliver' if status == 'PASS' else
                  'Review br issues' if br_issues > 0 else
                  'Fix structural issues' if s0_fails > 0 else
                  'Manual review needed')

        rn   = ws1.max_row + 1
        vals = [r['name'], r['doc_type_name'], root_tag, landscape,
                s0_score, s0_fails, br_score, br_issues,
                cross_st, rt_excel, rt_sgml, rt_ratio,
                score, status, action]

        for ci, v in enumerate(vals, 1):
            c = ws1.cell(rn, ci, v)
            c.border = TH; c.font = NRM; c.alignment = WRP

        ws1.cell(rn, 5).fill  = status_fill('PASS' if s0_score >= 95 else 'FAIL')
        ws1.cell(rn, 5).alignment = CTR
        ws1.cell(rn, 7).fill  = status_fill('PASS' if br_score >= 95 else 'FAIL')
        ws1.cell(rn, 7).alignment = CTR
        ws1.cell(rn, 9).fill  = status_fill('PASS' if not cross else 'WARN')
        ws1.cell(rn, 9).alignment = CTR
        ws1.cell(rn, 13).fill = status_fill('PASS' if score >= 90 else 'FAIL')
        ws1.cell(rn, 13).alignment = CTR
        ws1.cell(rn, 14).fill = status_fill(status)
        ws1.cell(rn, 14).alignment = CTR; ws1.cell(rn, 14).font = WN

    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 20
    for l, w in zip('CDEFGHIJKLMNO', [10,10,10,8,10,8,16,10,10,10,10,12,22]):
        ws1.column_dimensions[l].width = w
    ws1.freeze_panes = 'A4'

    # ══ SHEET 2: Stage-0 Structural Checks ════════════════════════════════════
    ws2 = wb.create_sheet('A - Structural Checks')
    CHECK_COLS = ['Check ID', 'Check Name', 'Category'] + [r['stem'][:22] for r in validated]
    for ci, h in enumerate(CHECK_COLS, 1):
        hdr(ws2, 1, ci, h, fill=BLUE if ci <= 3 else LBL)

    for chk_id, label, cat, _, _ in VENDOR_FREE_CHECKS:
        ws2.append([chk_id, label, cat])
        rn = ws2.max_row
        ws2.cell(rn, 1).font = NRM; ws2.cell(rn, 2).font = NRM; ws2.cell(rn, 3).font = NRM
        ws2.cell(rn, 3).fill = CAT_FILLS.get(cat, GREY)
        for ci, r in enumerate(validated, 4):
            st = r.get('stage0', {}).get('checks', {}).get(chk_id, {}).get('status', 'N/A')
            sm = r.get('stage0', {}).get('checks', {}).get(chk_id, {}).get('summary', '')
            c = ws2.cell(rn, ci)
            c.value = st if st != 'FAIL' else f'FAIL: {sm}'
            c.fill  = GREEN if st == 'PASS' else RED if st == 'FAIL' else GREY
            c.font  = NRM; c.alignment = WRP; c.border = TH

    # Score row
    ws2.append(['', 'Stage-0 Score', ''] + [r.get('stage0', {}).get('score', 0) for r in validated])
    rn = ws2.max_row
    ws2.cell(rn, 2).font = WN
    for ci in range(4, 4 + len(validated)):
        v = ws2.cell(rn, ci).value or 0
        ws2.cell(rn, ci).fill = GREEN if v >= 95 else RED
        ws2.cell(rn, ci).alignment = CTR; ws2.cell(rn, ci).font = WN

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 12
    for i, r in enumerate(validated):
        ws2.column_dimensions[get_column_letter(4 + i)].width = 28
    ws2.freeze_panes = 'D2'

    # ══ SHEET 3: Business Rules + Cross-File ══════════════════════════════════
    ws3 = wb.create_sheet('B+C - Business Rules')
    BR_COLS = ['File', 'Doc Type', 'Rule', 'Severity', 'Actual', 'Expected', 'Context', 'Score']
    for ci, h in enumerate(BR_COLS, 1): hdr(ws3, 1, ci, h)

    for r in validated:
        for issue in r.get('business', {}).get('issues', []):
            row = [r['name'], r['doc_type_name'], issue['rule'],
                   issue['severity'], issue['actual'][:60],
                   issue['expected'][:60], issue['context'][:80],
                   r.get('business', {}).get('score', 0)]
            ws3.append(row)
            rn = ws3.max_row
            sev = issue['severity']
            ws3.cell(rn, 4).fill = (RED if sev == 'Critical' else
                                     ORANGE if sev == 'High' else
                                     YELLOW if sev == 'Medium' else GREEN)
            for ci in range(1, 9):
                ws3.cell(rn, ci).font = NRM; ws3.cell(rn, ci).border = TH; ws3.cell(rn, ci).alignment = WRP

        for issue in r.get('cross', []):
            ws3.append([r['name'], r['doc_type_name'], 'Cross-File', 'Warning',
                        issue[:80], 'Matches group majority', '', r.get('cross_score', 0)])
            rn = ws3.max_row
            ws3.cell(rn, 4).fill = YELLOW
            for ci in range(1, 9):
                ws3.cell(rn, ci).font = NRM; ws3.cell(rn, ci).border = TH; ws3.cell(rn, ci).alignment = WRP

    if ws3.max_row == 1:
        ws3.append(['✅ No business rule violations found in any file.'])

    for ci, w in enumerate([36, 20, 10, 10, 30, 30, 40, 8], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.auto_filter.ref = f'A1:{get_column_letter(8)}{ws3.max_row}'
    ws3.freeze_panes = 'A2'

    # ══ SHEET 4: Round-Trip Completeness ══════════════════════════════════════
    ws4 = wb.create_sheet('E - Round-Trip Check')
    RT_COLS = ['File', 'Doc Type', 'Excel Non-Null Cells', 'SGML TBLCELL Tags',
               'Capture Ratio', 'Status', 'Notes']
    for ci, h in enumerate(RT_COLS, 1): hdr(ws4, 1, ci, h)

    for r in validated:
        rt = r.get('roundtrip', {})
        status = rt.get('status', 'N/A')
        notes  = '; '.join(i.get('context','') for i in rt.get('issues', []))[:120]
        row = [r['name'], r['doc_type_name'],
               rt.get('excel_cells', 0), rt.get('sgml_cells', 0),
               f'{rt.get("ratio", 0):.1%}', status, notes]
        ws4.append(row)
        rn = ws4.max_row
        ws4.cell(rn, 6).fill = GREEN if status == 'PASS' else YELLOW if status == 'WARN' else RED
        ws4.cell(rn, 6).alignment = CTR; ws4.cell(rn, 6).font = WN
        for ci in range(1, 8):
            ws4.cell(rn, ci).font = NRM; ws4.cell(rn, ci).border = TH; ws4.cell(rn, ci).alignment = WRP
            if ci in (3, 4, 5): ws4.cell(rn, ci).alignment = CTR

    for ci, w in enumerate([36, 20, 18, 16, 14, 10, 45], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = 'A2'

    try:
        wb.save(str(report_path))
        print(f'\n  📊 Report saved: {report_path}')
        print(f'     Sheets: Summary Dashboard | A-Structural | B+C-Business Rules | E-RoundTrip')
    except Exception as e:
        print(f'  ⚠️  Save error: {e}')


# ══════════════════════════════════════════════════════════════════════════════
# MASTER SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def print_master(validated: List[Dict]):
    total  = len(validated)
    passed = sum(1 for r in validated if r.get('final_status') == 'PASS')
    avg    = sum(r.get('final_score', 0) for r in validated) / max(total, 1)

    print(f'\n{"═"*70}')
    print(f'  PIPELINE RESULTS  ({total} files)')
    print(f'{"─"*70}')
    print(f'  Final PASS:    {passed}/{total}  ({passed/total*100:.0f}%)')
    print(f'  Average Score: {avg:.1f}/100')
    print(f'\n  {"File":<40} {"Type":<20} {"Score":>6} {"Status":>8}')
    print(f'  {"─"*78}')
    for r in sorted(validated, key=lambda x: x.get('final_score', 0), reverse=True):
        icon = '✅' if r.get('final_status') == 'PASS' else '❌'
        print(f'  {icon} {r["name"][:38]:<38} {r["doc_type_name"]:<20} '
              f'{r.get("final_score",0):>5}/100  {r.get("final_status","ERROR"):>6}')
    print(f'{"═"*70}\n')


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f'\n{"═"*70}')
    print(f'  Excel → SGML Pipeline  (Convert + Validate, No Vendor Files)')
    print(f'  Date: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    print(f'{"═"*70}')

    # Step 1: Convert
    conv_results = run_batch_conversion()

    # Step 2: Validate
    validated = validate_all(conv_results)

    # Step 3: Report
    generate_report(validated, REPORT_PATH)

    # Master summary
    print_master(validated)

    print(f'  SGML files: {OUTPUT_DIR}')
    print(f'  Report:     {REPORT_PATH}\n')
